"""Hoja de conteo en PDF y ciclo quincenal con Excel.

La hoja PDF es solo para revisar caminando el vivero: no ajusta nada. El
Excel sí termina en ajustes, pero siempre pasando por la pantalla de
diferencias y la confirmación del empleado.
"""

import io
import os
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection

from . import datos

ENCABEZADOS = ["Código", "Nombre", "Categoría", "En sistema", "Conteo físico"]
_LOGO = os.path.join(os.path.dirname(__file__), "static", "logo.jpg")


def _fecha_bonita(momento=None):
    momento = momento or datetime.now(datos.ZONA_PANAMA)
    return momento.strftime("%d/%m/%Y %I:%M %p").lower()


def _latin1(texto):
    """Las fuentes base del PDF solo saben latin-1: el guion largo y
    cualquier otro carácter fuera del alfabeto se sustituyen en vez de
    reventar con un nombre de producto raro."""
    return (texto or "").replace("—", "-").encode("latin-1", "replace").decode("latin-1")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

class _HojaConteo(FPDF):
    def __init__(self, fecha):
        # Carta, como el diseño base de los documentos de la compañía.
        super().__init__(orientation="P", unit="mm", format="Letter")
        self._fecha = fecha
        self.set_auto_page_break(auto=True, margin=18)

    def header(self):
        if os.path.exists(_LOGO):
            self.image(_LOGO, x=10, y=8, w=16)
        self.set_font("helvetica", "B", 14)
        self.set_xy(30, 9)
        self.cell(0, 7, _latin1("Hoja de conteo — Control de Stock"))
        self.set_font("helvetica", "", 9)
        self.set_xy(30, 16)
        self.set_text_color(90)
        self.cell(0, 5, _latin1(f"Plantas Panamá · Vivero Rose · {self._fecha}"))
        self.set_text_color(0)
        self.ln(14)
        self._fila_encabezado()

    def footer(self):
        self.set_y(-14)
        self.set_font("helvetica", "", 8)
        self.set_text_color(120)
        self.cell(0, 6, _latin1(f"Página {self.page_no()} — la hoja es de revisión, no ajusta nada"),
                  align="C")

    def _fila_encabezado(self):
        self.set_font("helvetica", "B", 9)
        self.set_fill_color(245, 239, 228)  # crema del prototipo
        for texto, ancho in zip(ENCABEZADOS, _ANCHOS):
            self.cell(ancho, 8, _latin1(texto), border=1, fill=True)
        self.ln()


_ANCHOS = [42, 68, 42, 22, 22]  # mm; suma < 216 - márgenes


def generar_pdf(inventario, ruta):
    """Escribe la hoja de conteo en `ruta`: todo el inventario ordenado por
    categoría y nombre, con la cantidad física del sistema y la columna de
    conteo en blanco."""
    pdf = _HojaConteo(_fecha_bonita())
    pdf.add_page()
    pdf.set_font("helvetica", "", 9)
    ordenados = sorted(inventario, key=lambda p: (p["categoria"], p["nombre"]))
    for producto in ordenados:
        for valor, ancho in zip(
            [producto["sku"], producto["nombre"], producto["categoria"],
             str(producto["fisico"]), ""],
            _ANCHOS,
        ):
            pdf.cell(ancho, 8, _latin1(valor), border=1)
        pdf.ln()
    pdf.output(ruta)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def plantilla_excel(inventario):
    """La plantilla del conteo quincenal, como bytes de .xlsx. Las columnas
    de identificación quedan protegidas; solo 'Conteo físico' es editable."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Conteo"
    hoja.append(ENCABEZADOS)
    for celda in hoja[1]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor="F5EFE4")
    for producto in sorted(inventario, key=lambda p: (p["categoria"], p["nombre"])):
        hoja.append([producto["sku"], producto["nombre"], producto["categoria"],
                     producto["fisico"], None])
    for columna, ancho in zip("ABCDE", [22, 34, 22, 12, 14]):
        hoja.column_dimensions[columna].width = ancho
    # Protección: la hoja entera bloqueada salvo la columna E (el conteo).
    for fila in hoja.iter_rows(min_row=2, min_col=5, max_col=5):
        for celda in fila:
            celda.protection = Protection(locked=False)
    hoja.protection.sheet = True
    salida = io.BytesIO()
    libro.save(salida)
    return salida.getvalue()


def leer_conteo_excel(contenido, inventario):
    """Lee el .xlsx llenado y lo valida contra el inventario actual.

    Devuelve (diferencias, sin_contar, errores):
    - diferencias: [{sku, nombre, en_sistema, contado}] solo donde el conteo
      difiere de lo físico del sistema HOY (en_sistema es el candado
      `esperada` que viaja al order-api).
    - sin_contar: cuántas filas quedaron con el conteo en blanco (se saltan).
    - errores: mensajes por fila inválida; con errores no se aplica nada.
    """
    por_sku = {p["sku"]: p for p in inventario}
    try:
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        return [], 0, ["El archivo no se pudo leer como .xlsx"]
    hoja = libro.active
    diferencias, sin_contar, errores, vistos = [], 0, [], set()
    for numero, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
        sku = str(fila[0]).strip() if fila and fila[0] is not None else ""
        if not sku:
            continue
        contado = fila[4] if len(fila) > 4 else None
        if sku not in por_sku:
            errores.append(f"Fila {numero}: el código {sku} no está en el inventario")
            continue
        if sku in vistos:
            errores.append(f"Fila {numero}: el código {sku} está repetido")
            continue
        vistos.add(sku)
        if contado is None or str(contado).strip() == "":
            sin_contar += 1
            continue
        try:
            contado = int(contado)
        except (TypeError, ValueError):
            errores.append(f"Fila {numero}: '{contado}' no es una cantidad")
            continue
        if contado < 0:
            errores.append(f"Fila {numero}: la cantidad no puede ser negativa")
            continue
        producto = por_sku[sku]
        if contado != producto["fisico"]:
            diferencias.append({
                "sku": sku,
                "nombre": producto["nombre"],
                "en_sistema": producto["fisico"],
                "contado": contado,
            })
    return diferencias, sin_contar, errores
