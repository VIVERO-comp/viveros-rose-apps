import io

from openpyxl import load_workbook

from app import conteos, datos
from tests_utiles import xlsx_de_conteo


def test_plantilla_trae_el_inventario_y_protege_columnas(con_inventario):
    contenido = conteos.plantilla_excel(con_inventario)
    hoja = load_workbook(io.BytesIO(contenido)).active
    filas = list(hoja.iter_rows(values_only=True))
    assert filas[0] == ("Código", "Nombre", "Categoría", "En sistema", "Conteo físico")
    # Ordenada por categoría y nombre; 'En sistema' es lo físico.
    assert [f[0] for f in filas[1:]] == ["PL-ALBAHACA", "PL-ROMERO", "PL-IXORA", "PL-PALMA"]
    assert filas[2][3] == 2 and filas[3][3] == 5
    assert hoja.protection.sheet is True
    # Solo la columna del conteo es editable.
    assert hoja["E2"].protection.locked is False
    assert hoja["A2"].protection.locked is True


def test_leer_conteo_detecta_diferencias_y_saltas(con_inventario):
    contenido = xlsx_de_conteo([
        ("PL-ROMERO", 6),        # difiere de 2
        ("PL-IXORA", 5),         # igual a lo físico: sin diferencia
        ("PL-PALMA", None),      # sin contar: se salta
    ])
    diferencias, sin_contar, errores = conteos.leer_conteo_excel(contenido, con_inventario)
    assert errores == []
    assert sin_contar == 1
    assert diferencias == [
        {"sku": "PL-ROMERO", "nombre": "Romero", "en_sistema": 2, "contado": 6},
    ]


def test_leer_conteo_valida_fila_por_fila(con_inventario):
    contenido = xlsx_de_conteo([
        ("PL-NADA", 4),
        ("PL-ROMERO", "tres"),
        ("PL-IXORA", -2),
        ("PL-ROMERO", 6),  # repetido
    ])
    diferencias, _, errores = conteos.leer_conteo_excel(contenido, con_inventario)
    assert len(errores) == 4
    assert "PL-NADA" in errores[0]
    assert "tres" in errores[1]
    assert "negativa" in errores[2]
    assert "repetido" in errores[3]


def test_archivo_ilegible_no_revienta(con_inventario):
    _, _, errores = conteos.leer_conteo_excel(b"esto no es un xlsx", con_inventario)
    assert errores == ["El archivo no se pudo leer como .xlsx"]


def test_pdf_se_genera(tmp_path, con_inventario):
    ruta = tmp_path / "hoja.pdf"
    conteos.generar_pdf(con_inventario, str(ruta))
    assert ruta.read_bytes().startswith(b"%PDF")


# ---------------------------------------------------------------------------
# Flujo completo por las rutas
# ---------------------------------------------------------------------------

def _importar(cliente, filas):
    return cliente.post(
        "/conteos/importar",
        files={"archivo": ("conteo.xlsx", xlsx_de_conteo(filas),
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )


def test_importar_revisar_confirmar(cliente, con_inventario, ajustes_registrados):
    r = _importar(cliente, [("PL-ROMERO", 6), ("PL-IXORA", 1)])
    assert r.status_code == 303
    n = int(r.headers["location"].split("/")[2])

    revision = cliente.get(f"/conteos/{n}/revisar")
    assert "2 diferencias" in revision.text
    assert "Confirmar y ajustar" in revision.text
    assert ajustes_registrados == []  # nada viaja a Odoo sin confirmar

    cliente.post(f"/conteos/{n}/confirmar", follow_redirects=False)
    assert ajustes_registrados == [{
        "ajustes": [
            {"sku": "PL-ROMERO", "cantidad": 6, "esperada": 2},
            {"sku": "PL-IXORA", "cantidad": 1, "esperada": 5},
        ],
        "empleado": "genesis", "motivo": "conteo_quincenal",
    }]
    conteo = datos.conteo(n)
    assert conteo["estado"] == "confirmado"
    assert [r_["resultado"] for r_ in conteo["datos"]["resultados"]] == ["aplicado", "aplicado"]
    # La pantalla muestra el resultado por producto.
    assert "Ajustado" in cliente.get(f"/conteos/{n}/revisar").text


def test_conteo_sin_diferencias_queda_confirmado_directo(cliente, con_inventario,
                                                         ajustes_registrados):
    r = _importar(cliente, [("PL-ROMERO", 2), ("PL-IXORA", 5)])
    n = int(r.headers["location"].split("/")[2])
    assert datos.conteo(n)["estado"] == "confirmado"
    assert ajustes_registrados == []  # no había nada que ajustar
    # Y el reloj quincenal del score se reinició.
    assert "conteo hecho hoy" in cliente.get("/").text


def test_importar_con_errores_no_crea_conteo(cliente, con_inventario):
    r = _importar(cliente, [("PL-NADA", 4)])
    assert r.status_code == 200
    assert "PL-NADA" in r.text
    assert datos.conteos_recientes() == []


def test_descartar_conteo(cliente, con_inventario, ajustes_registrados):
    r = _importar(cliente, [("PL-ROMERO", 6)])
    n = int(r.headers["location"].split("/")[2])
    cliente.post(f"/conteos/{n}/descartar", follow_redirects=False)
    assert datos.conteo(n)["estado"] == "descartado"
    # Un conteo descartado ya no se puede confirmar.
    cliente.post(f"/conteos/{n}/confirmar", follow_redirects=False)
    assert ajustes_registrados == []


def test_conflicto_al_confirmar_queda_visible(cliente, con_inventario, monkeypatch):
    def con_conflicto(ajustes, empleado, motivo):
        return {"ok": True, "resultados": [
            {"sku": a["sku"], "cantidad": a["cantidad"], "resultado": "conflicto",
             "anterior": 9}
            for a in ajustes
        ]}
    monkeypatch.setattr(datos, "ajustar_en_odoo", con_conflicto)
    r = _importar(cliente, [("PL-ROMERO", 6)])
    n = int(r.headers["location"].split("/")[2])
    cliente.post(f"/conteos/{n}/confirmar", follow_redirects=False)
    pantalla = cliente.get(f"/conteos/{n}/revisar").text
    assert "Cambió: hay 9" in pantalla
    assert "no se tocó" in pantalla
